import io
import json
import logging
import os
import tempfile
import time
from typing import List, Optional

from fastapi import FastAPI, File, Form, HTTPException, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse, StreamingResponse
from openai import OpenAI
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from pdf2image import convert_from_path
from pdfminer.high_level import extract_text as pdfminer_extract_text
from pydantic import BaseModel, Field, ValidationError
from pypdf import PdfReader
import pytesseract

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger("proforma-backend")

MAX_PAGES = int(os.getenv("MAX_PAGES", "60"))
MAX_FILE_MB = int(os.getenv("MAX_FILE_MB", "25"))
TOTAL_BYTES_LIMIT = MAX_FILE_MB * 1024 * 1024
MODEL_NAME = os.getenv("OPENAI_MODEL", "gpt-4.1-mini")


class ConsolidatedRow(BaseModel):
    account_number: Optional[str] = None
    account_name: str
    y2022: Optional[float] = None
    y2023: Optional[float] = None
    y2024: Optional[float] = None
    ttm: Optional[float] = None
    mapped_coa_code: Optional[str] = None
    mapped_coa_name: Optional[str] = None
    mapping_confidence: Optional[float] = Field(default=None, ge=0, le=1)
    confidence: Optional[float] = Field(default=None, ge=0, le=1)
    notes: Optional[str] = None


class ConsolidatedMeta(BaseModel):
    units: Optional[str] = None
    ttm_present: bool
    warnings: List[str] = Field(default_factory=list)


class ConsolidatedResponse(BaseModel):
    meta: ConsolidatedMeta
    rows: List[ConsolidatedRow]


app = FastAPI(title="Pro Forma Consolidator")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.get("/api/health")
def health() -> dict:
    return {"status": "ok"}


def extract_text_first(pdf_bytes: bytes) -> str:
    text_parts = []
    try:
        with io.BytesIO(pdf_bytes) as bio:
            reader = PdfReader(bio)
            for page in reader.pages[:MAX_PAGES]:
                text_parts.append(page.extract_text() or "")
    except Exception:
        logger.info("pypdf extraction failed; falling back to pdfminer")

    text = "\n".join(text_parts).strip()
    if len(text) > 100:
        return text

    try:
        with tempfile.NamedTemporaryFile(suffix=".pdf", delete=True) as tmp:
            tmp.write(pdf_bytes)
            tmp.flush()
            text = pdfminer_extract_text(tmp.name, maxpages=MAX_PAGES) or ""
            return text.strip()
    except Exception:
        return ""


def extract_with_ocr(pdf_bytes: bytes) -> str:
    text = ""
    with tempfile.NamedTemporaryFile(suffix=".pdf", delete=True) as tmp:
        tmp.write(pdf_bytes)
        tmp.flush()
        images = convert_from_path(tmp.name, first_page=1, last_page=MAX_PAGES)
        for image in images:
            text += "\n" + pytesseract.image_to_string(image)
    return text.strip()


def call_openai_for_consolidation(extracted_text: str, coa_csv: str, warnings: List[str]) -> ConsolidatedResponse:
    api_key = os.getenv("OPENAI_API_KEY")
    if not api_key:
        raise HTTPException(status_code=500, detail="OPENAI_API_KEY is not set")

    client = OpenAI(api_key=api_key)
    schema_hint = {
        "meta": {"units": "string|null", "ttm_present": True, "warnings": ["string"]},
        "rows": [
            {
                "account_number": "string|null",
                "account_name": "string",
                "y2022": 0,
                "y2023": 0,
                "y2024": 0,
                "ttm": 0,
                "mapped_coa_code": "string|null",
                "mapped_coa_name": "string|null",
                "mapping_confidence": 0.0,
                "confidence": 0.0,
                "notes": "string|null",
            }
        ],
    }

    prompt = f"""
You are a financial statement extraction engine.
Return JSON only.
Use this exact schema shape: {json.dumps(schema_hint)}

Rules:
- Never guess missing numbers; use null.
- Convert parentheses to negative numbers.
- Detect units (ones/thousands/millions), normalize values consistently, set meta.units.
- Only include TTM if explicitly present.
- Merge duplicates cautiously; if unsure keep separate and note.
- mapping_confidence and confidence should be 0..1.

COA CSV (may be empty):
{coa_csv or '(none)'}

Source text:
{extracted_text[:200000]}

Existing warnings from parsing pipeline:
{json.dumps(warnings)}
"""

    response = client.responses.create(
        model=MODEL_NAME,
        input=[{"role": "user", "content": [{"type": "text", "text": prompt}]}],
    )
    raw = response.output_text.strip()

    def parse_or_repair(raw_text: str) -> ConsolidatedResponse:
        try:
            payload = json.loads(raw_text)
            return ConsolidatedResponse.model_validate(payload)
        except (json.JSONDecodeError, ValidationError):
            repair = client.responses.create(
                model=MODEL_NAME,
                input=[
                    {
                        "role": "user",
                        "content": [
                            {
                                "type": "text",
                                "text": f"Return valid JSON only, no markdown, matching the schema exactly. Repair this:\n{raw_text}",
                            }
                        ],
                    }
                ],
            )
            payload = json.loads(repair.output_text.strip())
            return ConsolidatedResponse.model_validate(payload)

    return parse_or_repair(raw)


@app.post("/api/consolidate", response_model=ConsolidatedResponse)
async def consolidate(pdfs: List[UploadFile] = File(...), coa_csv: str = Form(default="")):
    started = time.time()
    if not pdfs:
        raise HTTPException(status_code=400, detail="No PDFs uploaded")

    warnings: List[str] = []
    combined_text_parts: List[str] = []
    total_bytes = 0

    for file in pdfs:
        if file.content_type != "application/pdf":
            warnings.append(f"Skipped non-PDF file: {file.filename}")
            continue

        raw = await file.read()
        total_bytes += len(raw)
        if total_bytes > TOTAL_BYTES_LIMIT:
            raise HTTPException(status_code=413, detail=f"Total file size exceeds {MAX_FILE_MB}MB")

        txt = extract_text_first(raw)
        if len(txt) < 300:
            warnings.append(f"Scanned or image-heavy PDF detected for {file.filename}; OCR used")
            try:
                txt = extract_with_ocr(raw)
            except Exception:
                warnings.append(f"OCR failed for {file.filename}")
                continue

        if not txt.strip():
            warnings.append(f"No text extracted from {file.filename}")
            continue

        combined_text_parts.append(f"\n--- FILE: {file.filename} ---\n{txt}")

    if not combined_text_parts:
        raise HTTPException(status_code=400, detail="Could not extract text from any uploaded PDF")

    joined = "\n".join(combined_text_parts)
    result = call_openai_for_consolidation(joined, coa_csv, warnings)
    result.meta.warnings.extend([w for w in warnings if w not in result.meta.warnings])

    elapsed = round(time.time() - started, 2)
    logger.info("consolidate finished files=%s total_bytes=%s elapsed=%ss", len(pdfs), total_bytes, elapsed)
    return JSONResponse(content=result.model_dump())


@app.post("/api/export-xlsx")
def export_xlsx(payload: ConsolidatedResponse):
    wb = Workbook()
    ws = wb.active
    ws.title = "Consolidated"

    headers = [
        "Account Number",
        "Account Name",
        "2022",
        "2023",
        "2024",
        "TTM",
        "Mapped COA Code",
        "Mapped COA Name",
        "Mapping Confidence",
        "Confidence",
        "Notes",
    ]
    ws.append(headers)

    for cell in ws[1]:
        cell.font = Font(bold=True)

    for row in payload.rows:
        ws.append([
            row.account_number,
            row.account_name,
            row.y2022,
            row.y2023,
            row.y2024,
            row.ttm,
            row.mapped_coa_code,
            row.mapped_coa_name,
            row.mapping_confidence,
            row.confidence,
            row.notes,
        ])

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    widths = [18, 30, 12, 12, 12, 12, 18, 24, 18, 12, 30]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    for row in ws.iter_rows(min_row=2, min_col=3, max_col=6):
        for cell in row:
            cell.number_format = '#,##0.00_);(#,##0.00)'

    for row in ws.iter_rows(min_row=2, min_col=9, max_col=10):
        for cell in row:
            cell.number_format = '0.00'

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="consolidated.xlsx"'},
    )
